"""Generate bilingual XLSX templates for Firmyx file upload."""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.comments import Comment

OUTPUT_DIR = os.path.join(os.path.dirname(__file__), "..", "frontend", "public")

MONTHLY_COLUMNS = [
    "month", "year", "revenue", "expenses", "payroll", "rent",
    "debt", "cash_reserves", "taxes", "cogs",
    "total_assets", "current_liabilities", "ebit", "retained_earnings",
]

ANNUAL_COLUMNS = [
    "year", "revenue", "expenses", "payroll", "rent",
    "debt", "cash_reserves", "taxes", "cogs",
    "total_assets", "current_liabilities", "ebit", "retained_earnings",
]

MONTHLY_SAMPLE_ROWS = [
    [1, 2026, 62000, 43000, 18000, 3500, 12000, 25000, 3100, 15000, 175000, 42000, 11000, 35000],
    [2, 2026, 65000, 44500, 18500, 3500, 11500, 27000, 3250, 15500, 178000, 41000, 12500, 37000],
    [3, 2026, 68000, 46000, 19000, 3500, 11000, 29500, 3400, 16000, 182000, 40000, 13500, 39000],
]

ANNUAL_SAMPLE_ROWS = [
    [2025, 750000, 520000, 210000, 42000, 12000, 25000, 38000, 180000, 175000, 42000, 120000, 35000],
]

COLUMN_DESCRIPTIONS_EN = {
    "month": "Month number (1–12)",
    "year": "Calendar year (e.g. 2026)",
    "revenue": "Total monthly/annual revenue (sales income)",
    "expenses": "Total monthly/annual operating expenses",
    "payroll": "Total salaries and wages paid",
    "rent": "Rent and lease payments",
    "debt": "Total outstanding debt (loans, credit lines)",
    "cash_reserves": "Available cash and liquid assets",
    "taxes": "Tax payments for the period",
    "cogs": "Cost of goods sold (direct production costs)",
    "total_assets": "Total value of all assets",
    "current_liabilities": "Short-term obligations due within 12 months",
    "ebit": "Earnings before interest and taxes",
    "retained_earnings": "Cumulative net income retained in the business",
}

COLUMN_DESCRIPTIONS_BG = {
    "month": "Номер на месец (1–12)",
    "year": "Календарна година (напр. 2026)",
    "revenue": "Общи месечни/годишни приходи (продажби)",
    "expenses": "Общи месечни/годишни оперативни разходи",
    "payroll": "Общо изплатени заплати и възнаграждения",
    "rent": "Наем и лизингови плащания",
    "debt": "Общ непогасен дълг (заеми, кредитни линии)",
    "cash_reserves": "Налични парични средства и ликвидни активи",
    "taxes": "Данъчни плащания за периода",
    "cogs": "Себестойност на продадените стоки (преки производствени разходи)",
    "total_assets": "Обща стойност на всички активи",
    "current_liabilities": "Краткосрочни задължения с падеж до 12 месеца",
    "ebit": "Печалба преди лихви и данъци",
    "retained_earnings": "Натрупана неразпределена печалба",
}

HEADER_FILL = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
DATA_FONT = Font(size=11)


def style_header(ws, columns):
    """Apply header styling to the first row."""
    for col_idx, col_name in enumerate(columns, 1):
        cell = ws.cell(row=1, column=col_idx, value=col_name)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = Alignment(horizontal="center")
        ws.column_dimensions[cell.column_letter].width = max(len(col_name) + 4, 14)


def add_data_rows(ws, rows, start_row=2):
    """Write sample data rows."""
    for row_idx, row_data in enumerate(rows, start_row):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = Alignment(horizontal="center")


def add_instructions_sheet(wb, columns, descriptions, title="Instructions"):
    """Add an instructions sheet explaining each column."""
    ws = wb.create_sheet(title=title)
    ws.cell(row=1, column=1, value="Column").font = Font(bold=True, size=11)
    ws.cell(row=1, column=2, value="Description" if title == "Instructions" else "Описание").font = Font(bold=True, size=11)
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 60
    for idx, col_name in enumerate(columns, 2):
        ws.cell(row=idx, column=1, value=col_name).font = DATA_FONT
        ws.cell(row=idx, column=2, value=descriptions.get(col_name, "")).font = DATA_FONT


def add_bg_comments(ws, columns, descriptions):
    """Add Bulgarian cell comments to header row."""
    for col_idx, col_name in enumerate(columns, 1):
        desc = descriptions.get(col_name, "")
        if desc:
            ws.cell(row=1, column=col_idx).comment = Comment(desc, "Firmyx")


def generate_monthly_en():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    style_header(ws, MONTHLY_COLUMNS)
    add_data_rows(ws, MONTHLY_SAMPLE_ROWS)
    add_instructions_sheet(wb, MONTHLY_COLUMNS, COLUMN_DESCRIPTIONS_EN, "Instructions")
    path = os.path.join(OUTPUT_DIR, "template_monthly_en.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def generate_monthly_bg():
    wb = Workbook()
    ws = wb.active
    ws.title = "Данни"
    style_header(ws, MONTHLY_COLUMNS)
    add_data_rows(ws, MONTHLY_SAMPLE_ROWS)
    add_bg_comments(ws, MONTHLY_COLUMNS, COLUMN_DESCRIPTIONS_BG)
    add_instructions_sheet(wb, MONTHLY_COLUMNS, COLUMN_DESCRIPTIONS_BG, "Инструкции")
    path = os.path.join(OUTPUT_DIR, "template_monthly_bg.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def generate_annual_en():
    wb = Workbook()
    ws = wb.active
    ws.title = "Data"
    style_header(ws, ANNUAL_COLUMNS)
    add_data_rows(ws, ANNUAL_SAMPLE_ROWS)
    add_instructions_sheet(wb, ANNUAL_COLUMNS, COLUMN_DESCRIPTIONS_EN, "Instructions")
    path = os.path.join(OUTPUT_DIR, "template_annual_en.xlsx")
    wb.save(path)
    print(f"Created: {path}")


def generate_annual_bg():
    wb = Workbook()
    ws = wb.active
    ws.title = "Данни"
    style_header(ws, ANNUAL_COLUMNS)
    add_data_rows(ws, ANNUAL_SAMPLE_ROWS)
    add_bg_comments(ws, ANNUAL_COLUMNS, COLUMN_DESCRIPTIONS_BG)
    add_instructions_sheet(wb, ANNUAL_COLUMNS, COLUMN_DESCRIPTIONS_BG, "Инструкции")
    path = os.path.join(OUTPUT_DIR, "template_annual_bg.xlsx")
    wb.save(path)
    print(f"Created: {path}")


if __name__ == "__main__":
    os.makedirs(OUTPUT_DIR, exist_ok=True)
    generate_monthly_en()
    generate_monthly_bg()
    generate_annual_en()
    generate_annual_bg()
    print("All templates generated successfully.")
